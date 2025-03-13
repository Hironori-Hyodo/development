
from fastapi import APIRouter,File,Form,UploadFile,Request,HTTPException
from fastapi.responses import JSONResponse,Response
import openpyxl as op
from io import BytesIO
# from models import init_db
from datetime import datetime
import os
import sqlite3
import subprocess

router = APIRouter()

DB_FILE = "excel_data.db"
UPLOAD_DIR = "temp"
os.makedirs(UPLOAD_DIR, exist_ok=True)


# DB 初期化（テーブルを削除して再作成）
def init_db():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # **🔴 既存のテーブルを削除**
    cursor.execute("DROP TABLE IF EXISTS screenshots")
    
    # **🟢 新しいテーブルを作成**
    cursor.execute("""
        CREATE TABLE screenshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_name TEXT,
            update_date TEXT,
            image BLOB
        )
    """)
    
    conn.commit()
    conn.close()

# **FastAPI 起動時にテーブルをリセット**
init_db()

# 画像を SQLite に保存
def save_screenshot_to_db(sheet_name, update_date, image_bytes):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO screenshots (sheet_name, update_date, image) VALUES (?, ?, ?)",
                   (sheet_name, update_date, image_bytes))
    conn.commit()
    conn.close()


# Excel の特定シートを PDF → 画像（PNG） に変換
def convert_excel_sheet_to_image(excel_path, sheet_name):
    sheet_pdf_path = os.path.join(UPLOAD_DIR, f"{sheet_name}.pdf")
    sheet_image_path = os.path.join(UPLOAD_DIR, f"{sheet_name}.png")

    # LibreOffice を使って Excel を PDF に変換
    subprocess.run([
        "xvfb-run", "--server-args=-screen 0 1024x768x24", "libreoffice", "--headless",
        "--convert-to", "pdf", "--outdir", UPLOAD_DIR, excel_path
    ], check=True)

    # PDF を PNG に変換
    subprocess.run(["pdftoppm", "-png", "-singlefile", sheet_pdf_path, sheet_image_path.replace(".png", "")], check=True)

    # PNG のバイナリデータを取得
    with open(sheet_image_path, "rb") as img_file:
        return img_file.read()



@router.post("/upload")
async def upload(files:UploadFile):
    print("Nice Upload! Post")

    init_db()

    # allow mime list
    ALLOWED_MIME_TYPES = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
        "application/vnd.ms-excel.sheet.macroEnabled.12",  # .xlsm
        "application/vnd.ms-excel.sheet.binary.macroEnabled.12",  # .xlsb
        "application/vnd.openxmlformats-officedocument.spreadsheetml.template",  # .xltx
        "application/vnd.ms-excel",  # .xlt (Excel 97-2003 Template)
        "application/vnd.ms-excel.addin.macroEnabled.12",  # .xlam
    ]

    # Readable only Excel file
    if files.content_type not in ALLOWED_MIME_TYPES:
        raise HTTPException(status_code=400,detail="無効なファイル形式です。\nExcelファイルをアップロードしてください。")


    workbook_name = files.filename
    contents = await files.read()


    print("Excel is save!")
    excel_path = os.path.join(UPLOAD_DIR, workbook_name)
    with open(excel_path, "wb") as f:
        f.write(contents)

    print("E",excel_path)

    # Get current date
    dt = datetime.now()
    datetime_str = dt.strftime("%Y.%m.%d")

    # print("Get a screen-shot!")
    # # スクリーンショット撮影
    # image_path = excel_path.replace(".xlsx", ".png")
    # image_bytes = convert_excel_to_image(excel_path, image_path)


    # print("Save DB!")
    # # DB に保存
    # save_screenshot_to_db(workbook_name, datetime_str, image_bytes)


    try:
        workbook = op.load_workbook(BytesIO(contents))
    except Exception as e:
        print(e)
        raise HTTPException(status_code=400,detail="Excelファイルの読み込みに失敗しました。\nファイルが破損している可能性があります。\nシステム管理者にご連絡ください。")


    sheets_data = []
    for sheet_name in workbook.sheetnames:

        print(f"Processing sheet: {sheet_name}")

        # **各シートのスクリーンショットを撮影**
        image_bytes = convert_excel_sheet_to_image(excel_path, sheet_name)

        # **シートごとに DB に保存**
        save_screenshot_to_db(sheet_name, datetime_str, image_bytes)

        # # 左上(min-row * min-colomn)と右下(max-row * max-column)を取得
        # ws = workbook[sheet_name]

        # not_empty_cells = [cell for row in ws.iter_rows() for cell in row if cell.value is not None]
        # # print(not_empty_cells)

        # if not_empty_cells:
        #     min_row = min([cell.row for cell in not_empty_cells])
        #     max_row = max([cell.row for cell in not_empty_cells])
        #     min_column = min([cell.column for cell in not_empty_cells])
        #     max_column = max([cell.column for cell in not_empty_cells])

        #     for merged_range in ws.merged_cells:
        #         # print(merged_range)
        #         minr,minc,maxr,maxc = merged_range.min_row,merged_range.min_col,merged_range.max_row,merged_range.max_col
        #         max_row = max(max_row, maxr)
        #         max_column = max(max_column, maxc)

        sheets_data.append({
            "sheet_name":sheet_name,
            "update_date":datetime_str
        })

    
    return JSONResponse({"sheets":sheets_data})



# 画像を取得する API
@router.get("/screenshot/{sheet_name}")
async def get_screenshot(sheet_name: str):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT image FROM screenshots WHERE sheet_name=?", (sheet_name,))
    result = cursor.fetchone()
    conn.close()

    if result:
        return Response(content=result[0], media_type="image/png")
    else:
        raise HTTPException(status_code=404, detail="スクリーンショットが見つかりません")